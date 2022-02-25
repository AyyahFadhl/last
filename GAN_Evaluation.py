

import matplotlib.pyplot as plt
import tensorflow as tf
import numpy as np
import matlab.engine
import skimage
import pickle
import pandas as pd
import openpyxl
from xlsxwriter import Workbook
import os

''''-----------------------------------------------------Helping Functions------------------------------------------------------------'''
MasterPath = os.path.dirname(os.path.abspath(__file__))

'''Return image data from a raw PGM file as numpy array.
Format specification: http://netpbm.sourceforge.net/doc/pgm.html'''
#input: file name
#output: 4D tensorflow of image
def read_pgm(filename):
    image = plt.imread(filename,format='pgm')
    image= tf.reshape(tf.convert_to_tensor(image),(1,512,512,1))
    #Plotting read image
    #pyplot.imshow(image, pyplot.cm.gray)
    #pyplot.show()
    return image


''' Write Numpy array in PGM format to file.'''
#input: 4D tesnor of image, file name
#output:no output
def write_pgm(image, filename):
    #Converting Tensor to  2D numpy array
    image= tf.cast(image,dtype= tf.uint8)
    image=np.reshape(image.numpy(),(image.shape[1],image.shape[2]))
    height= image.shape[0]
    width = image.shape[1]
    maxval = image.max()
    with open(filename, 'wb') as f:
        f.write(bytes('P5\n{} {}\n{}\n'.format(width, height, maxval), 'ascii'))
        # not sure if next line works universally, but seems to work on my mac
        image.tofile(f)
    return

'''
Produce a randomly generated matrix (CoverImage H,CoverImage Width) of  zeros and once (Uniform distribution) using Numpy random (PRNG)'''
# input: ImageSize
# output: Randomly generated matrix of zeros and once. (Numpy type, float32)
def BinaryMatrixGenerator(ProbabilityMapSize):
    # initializing the seed for PRNG based on system (Truly RNG)
    np.random.seed(list(os.urandom(1)))

    # generating  binary numpy array of size 512x512
    Random_matrix = np.random.randint(low=0, high=2, size=ProbabilityMapSize)

    # plotting matrix
    #cmap = ListedColormap(['k', 'w'])
    #cax = plt.matshow(Random_matrix, cmap=cmap)
    #plt.show()

    # Casting to float32
    Random_matrix = np.cast['float32'](Random_matrix)
    return Random_matrix


'''
call the  BinaryMatrixGenerator Function to obtain BinaryMatrix of Numpy float32 type,
input the BinaryMatrix and probability map element-wise to the activation function using following equations
m_(i,j)= -0.5 tanh(? (p_(i,j)-2n_(i,j)))+ 0.5 tanh(? (p_(i,j)-2?(1-n?_(i,j)))).
tanh(x)=(e^x- e^(-x))/(e^x+ e^(-x) )'''
# input : Probability map p. (Numpy type, float32)
# output: Modification map m. (Numpy type,  float32)
def TernaryEmbeddingSimulator(p):
    # Call BinaryMatrixGenerator() to obtain n
    n = BinaryMatrixGenerator(p.shape)
    # Creating and computing  modificarionMap numpy array.
    #m = np.empty(p.shape, np.float32)
    # 1000 is the scalling factor to simulate stair case function
    m = -0.5 * tf.keras.backend.tanh(1000 * (p - 2 * n)) + 0.5 * tf.keras.backend.tanh(1000 * (p - 2 * (1 - n)))
    # Compute modificationMap
    # 1000 is the scalling factor to simulate stair case function

    '''
    for batch in range(p.shape[0]):
        for i in range(p.shape[1]):
            for j in range(p.shape[2]):
                m[batch,i, j,0] = -0.5 * tf.keras.backend.tanh(1000 * (p[batch,i, j,0] - 2 * n[batch,i, j,0])) + 0.5 * tf.keras.backend.tanh(1000 * (p[batch,i, j,0] - 2 * (1 - n[batch,i, j,0])))
    '''


    # print modificationMap
    # np.set_printoptions(threshold=np.inf)
    # print(m)

    # plotting modification matrix
    # plt.matshow(m)
    # plt.show()
    return m



''''--------------------------------------------------Set Up------------------------------------------------------------------------------------------'''
'''Generate Stego image using S_UNIWARD Algorithm for certain cover image and payload '''
#input: CoverImagePath and Payload
#output: Stego image
def GeneratingStegoImageUsingS_UNIWARDAlgorithm(CoverImage,Payload):
    # generating Stego image
    eng = matlab.engine.start_matlab()
    StegoImage = eng.S_UNIWARD(CoverImage, Payload)  # type mlarray double (matlab array double)
    StegoImage = np.array(StegoImage)  # change to numpy array of type float
    StegoImage = StegoImage.astype(np.uint8)  # change to unsigned int
    eng.quit()
    #Convert Numpy to 4D tensor
    StegoImage= tf.reshape(tf.convert_to_tensor(StegoImage), (1, 256, 256, 1))
    return StegoImage

'''Randomly partition the BOSSBase Dataset into training and testing 50% to train the  SRM_EC Steganalyzer'''
def Steganalayzer_Training_Testing_Data():
    os.makedirs(os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/CoverImages'), exist_ok=True)
    os.makedirs(os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/StegoImages'), exist_ok=True)
    os.makedirs(os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/CoverImages'), exist_ok=True)
    os.makedirs(os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/StegoImages'), exist_ok=True)
    # Get 5000 random permutation, which represents the Training cover images  from BOSSBase, resize them and generate stego images, save in the folder
    Dataset = np.arange(1, 10001)
    Dateset = np.random.permutation(Dataset)

    # Steganalyzer Training Data
    x=1
    for img_num in range(5000):
        try:
            # reading,resizing and saving  Cover image
            CoverImage_path = os.path.join(MasterPath, f'Dataset_Steganalyzer\BOSSbase_1.01\{Dateset[img_num]}.pgm')
            CoverImage = read_pgm(CoverImage_path)
            CoverImage = tf.image.resize(CoverImage, (256, 256))
            write_pgm(CoverImage,os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/CoverImages\{x}.pgm'))
            # Generating Stego image and savinng it
            StegoImage = GeneratingStegoImageUsingS_UNIWARDAlgorithm(os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/CoverImages\{x}.pgm'), 0.4)
            write_pgm(StegoImage,os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/StegoImages\{x}.pgm'))
        except Exception as e:
            print(img_num, "   ", Dateset[img_num])
            pass
        #Updating the counter
        x=x+1
    #Steganalyzer Testing Data
    x=1
    for img_num in range(5,10001):
        try:
            # reading,resizing and saving  Cover image
            CoverImage_path = os.path.join(MasterPath, f'Dataset_Steganalyzer\BOSSbase_1.01\{Dateset[img_num]}.pgm')
            CoverImage = read_pgm(CoverImage_path)
            CoverImage = tf.image.resize(CoverImage, (256, 256))
            write_pgm(CoverImage,os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/CoverImages\{x}.pgm'))
            # Generating Stego image and savinng it
            StegoImage = GeneratingStegoImageUsingS_UNIWARDAlgorithm(os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/CoverImages\{x}.pgm'),0.4)
            write_pgm(StegoImage, os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/StegoImages\{x}.pgm'))
        except Exception as e:
            print(img_num, "   ", Dateset[img_num])
            pass
            # Updating the counter
        x = x + 1
    return

def SRM_EC_setup():
    eng = matlab.engine.start_matlab()

    #SRM Feature extraction for Training Cover Image
    for cover in range(1,5001):
        image = os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/CoverImages/{cover}.pgm')
        SRMFeatures = eng.SRM(image)  # class dict
        SRMFeatures = np.array(list(SRMFeatures.items()) ) # convert dict to list, then convert to np array(2,106)
        SRMFeatures = SRMFeatures.transpose()  # transpose it to make it  (2, 106), instead of (106,2)
        # Now we have 2 row, and 106 columns (firstrow (0,:) have the name of the features)
        # row(1,:) have the matrix of the features  (in mlarray.single() formate)
        # mlarray: matlab array of Single formate value (Single is floating point values stored in 4 byte(32 bit), of data type class Single)
        # to display them in a readable manner
        #for j in range(106): print(SRMFeatures[0, j],' : ',SRMFeatures[1, j])
        #to display the shape
        #print(SRMFeatures.shape)  # 2, 106
        SRMFeatures = np.delete(SRMFeatures, (0), axis=0) # Remove the first row in SRMFeature
        if (cover==1):
            TrainingCoverImage_SRMFeatures=SRMFeatures
        else:
            TrainingCoverImage_SRMFeatures = np.vstack((TrainingCoverImage_SRMFeatures,SRMFeatures))  # add the computed SRM features to the next row in the Cover features SRM features.
    TrainingCoverImage_SRMFeatures = matlab.double(TrainingCoverImage_SRMFeatures.tolist())  # change to mlarray double, so we can send to ensemble Training.
        #now the shape of the CoverImage_SRMFeatures (#number of cover images,106), and the type is mlaraay.double


    #SRM Feature extraction for Training Stego Image
    for stego in range(1,5001):
        image = os.path.join(MasterPath, f'Dataset_Steganalyzer/TrainingSet/TrainingData256/StegoImages/{stego}.pgm')
        SRMFeatures = eng.SRM(image)  # class dict
        SRMFeatures = np.array(list(SRMFeatures.items()) ) # convert dict to list, then to numpy array
        SRMFeatures = SRMFeatures.transpose()  # transpose it to make it  (2, 106), instead of (106,2)
        # Now we have 2 row, and 106 columns (firstrow (0,:) have the name of the features)
        # row(1,:) have the matrix of the features  (in mlarray.single() formate)
        # mlarray: matlab array of Single formate value (Single is floating point values stored in 4 byte(32 bit), of data type class Single)
        # to display them in a readable manner
        # for j in range(106): print(SRMFeatures[0, j],' : ',SRMFeatures[1, j])
        #To display the shape
        #print(SRMFeatures.shape)  # 2, 106
        SRMFeatures = np.delete(SRMFeatures, (0), axis=0) # Remove the first row in SRMFeature
        if (stego==1):
            TrainingStegoImage_SRMFeatures=SRMFeatures
        else:
            TrainingStegoImage_SRMFeatures = np.vstack((TrainingStegoImage_SRMFeatures, SRMFeatures))  # add the computed SRM features to the next
    TrainingStegoImage_SRMFeatures = matlab.double(TrainingStegoImage_SRMFeatures.tolist())  # change to mlarray double, so we can send to ensemble Training.
    #now the shape of the CoverImage_SRMFeatures (#number of cover images,106), and the type is mlaraay.double


    #Train Ensemble Classifier
    TrainedEnsemble =eng.ensemble_training(TrainingCoverImage_SRMFeatures,TrainingStegoImage_SRMFeatures)

    # SRM Feature extraction for Testing Cover Image
    for cover in range(4996, 9996):
        image = os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/CoverImages/{cover}.pgm')
        SRMFeatures = eng.SRM(image)  # class dict
        SRMFeatures = np.array(list(SRMFeatures.items()))  # convert dict to list, then convert to np array(2,106)
        SRMFeatures = SRMFeatures.transpose()  # transpose it to make it  (2, 106), instead of (106,2)
        # Now we have 2 row, and 106 columns (firstrow (0,:) have the name of the features)
        # row(1,:) have the matrix of the features  (in mlarray.single() formate)
        # mlarray: matlab array of Single formate value (Single is floating point values stored in 4 byte(32 bit), of data type class Single)
        # to display them in a readable manner
        # for j in range(106): print(SRMFeatures[0, j],' : ',SRMFeatures[1, j])
        # to display the shape
        # print(SRMFeatures.shape)  # 2, 106
        SRMFeatures = np.delete(SRMFeatures, (0), axis=0)  # Remove the first row in SRMFeature
        if (cover == 4996):
            CoverImage_SRMFeatures = SRMFeatures
        else:
            CoverImage_SRMFeatures = np.vstack((CoverImage_SRMFeatures ,
                                                SRMFeatures))  # add the computed SRM features to the next row in the Cover features SRM features.
    TestingCoverImage_SRMFeatures = matlab.double(CoverImage_SRMFeatures .tolist())  # change to mlarray double, so we can send to ensemble Training.
    # now the shape of the CoverImage_SRMFeatures (#number of cover images,106), and the type is mlaraay.double

    # SRM Feature extraction for Testing Stego Image
    for stego in range(4996, 9996):
        image = os.path.join(MasterPath, f'Dataset_Steganalyzer/TestingSet/TestingData256/StegoImages/{stego}.pgm')
        SRMFeatures = eng.SRM(image)  # class dict
        SRMFeatures = np.array(list(SRMFeatures.items()))  # convert dict to list, then to numpy array
        SRMFeatures = SRMFeatures.transpose()  # transpose it to make it  (2, 106), instead of (106,2)
        # Now we have 2 row, and 106 columns (firstrow (0,:) have the name of the features)
        # row(1,:) have the matrix of the features  (in mlarray.single() formate)
        # mlarray: matlab array of Single formate value (Single is floating point values stored in 4 byte(32 bit), of data type class Single)
        # to display them in a readable manner
        # for j in range(106): print(SRMFeatures[0, j],' : ',SRMFeatures[1, j])
        # To display the shape
        # print(SRMFeatures.shape)  # 2, 106
        SRMFeatures = np.delete(SRMFeatures, (0), axis=0)  # Remove the first row in SRMFeature
        if (stego == 4996):
            StegoImage_SRMFeatures = SRMFeatures
        else:
            StegoImage_SRMFeatures = np.vstack((StegoImage_SRMFeatures , SRMFeatures))  # add the computed SRM features to the next
    TestinngStegoImage_SRMFeatures = matlab.double(StegoImage_SRMFeatures .tolist())  # change to mlarray double, so we can send to ensemble Training.
    # now the shape of the CoverImage_SRMFeatures (#number of cover images,106), and the type is mlaraay.double

    #Testing ensemble classifier (Testing on cover and stego seperately)
    TestingCoverresults=eng.ensemble_testing(TestingCoverImage_SRMFeatures,TrainedEnsemble)
    TestingStegoresults= eng.ensemble_testing(TestinngStegoImage_SRMFeatures,TrainedEnsemble)

    #computing the false alarm in testing cover and the missed detection in testing stego
    false_alarm=np.sum(np.array(TestingCoverresults["predictions"]) != -1.0)
    missed_detection=np.sum(np.array(TestingStegoresults["predictions"] )!= +1.0)
    Testing_Error=(false_alarm+missed_detection)/(np.shape(StegoImage_SRMFeatures)[0]+ np.shape(CoverImage_SRMFeatures )[0])

    Results=("1. False alarm in Cover",false_alarm, "\n2. Missed detection on stego ",missed_detection,"\n 3.Testing_Error",Testing_Error )
    print("1. False alarm in Cover",false_alarm, "\n2. Missed detection on stego ",missed_detection,"\n3.Testing_Error",Testing_Error)

    # Saving results of Testing SRM+EC
    with open(os.path.join(MasterPath,"Dataset_Steganalyzer/Result of Testing SRM+EC"), 'w') as f:
        f.write(str(Results))
        f.close()

    #Saving the trained ensemble
    with open(os.path.join(MasterPath,"Dataset_Steganalyzer/SRMTrainedEnsemble.pkl"), 'wb') as f:
        pickle.dump(TrainedEnsemble,f)
        f.close()

    return

def GBRAS_Net_setup():
    #manulally downloaded the model from github
    return


'''------------------------------------------------------------------------------------------------------------------'''

#input: path of the image
#output: prediction of the SRM_EC -1.0 Cover, +1.0 Stego
def SRM_EC(Image):
    eng = matlab.engine.start_matlab()
    #Generate the SRM feature map, reshape it and prepare it to be tested by trained ensemble
    SRMFeatures= eng.SRM(Image)
    SRMFeatures = np.array(list(SRMFeatures.items()))  # convert dict to list, then convert to np array(2,106)
    SRMFeatures = SRMFeatures.transpose()  # transpose it to make it  (2, 106), instead of (106,2)
    SRMFeatures = np.delete(SRMFeatures, (0), axis=0)

    # 1. load the trained SRM+EC model as list
    with open(os.path.join(MasterPath, "Dataset_Steganalyzer/SRMTrainedEnsemble.pkl"), 'rb') as f:
        TrainedEnsemble= pickle.load(f)
        f.close()

    # 2. call ensemble Testing to get the prediction, and save all the prediction in list to compute detection error later on
    Results= eng.ensemble_testing(matlab.double(SRMFeatures.tolist()), TrainedEnsemble)
    return Results["predictions"] # this return -1 cover or +1 stego



#input : 4D tensor
#output: Prediction of GBRAS_Net ensemble
def GBRAS_Net(Image):
        # loading the model
        model= tf.keras.models.load_model(os.path.join(MasterPath,f'Dataset_Steganalyzer/GBRAS_Net Trained on 0.4 S_UNIWARD.hdf5'),custom_objects={'Tanh3': Tanh3})
        #Image Prediction GBRAS_Net Trained on 0.4  S_UNIWARD
        oacc_val = model.predict(Image)
        prediction = np.round(oacc_val)
        if (prediction == np.array([1, 0])).all():
            return -1.0
        else:
            return +1.0

def Tanh3(x):
    tanh3 = tf.keras.activations.tanh(x)*3
    return tanh3


'''--------------------------------------------Main Function-----------------------------------------------------------------'''

def QuantitativeEvaluation(epoch, generator, TestX,epochTrainingTime):
      # Save the images generated during each epoch.
      os.makedirs(os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}'), exist_ok=True)
      os.makedirs(os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_CoverImages'), exist_ok=True)
      os.makedirs(os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_StegoImages'),exist_ok=True)
      np.random.shuffle(TestX)
      mini_Batch_TestX = np.array_split(TestX, TestX.shape[0])
      SRMPredictionC= []
      SRMPredictionS = []
      GBRAS_NetPredictionC=[]
      GBRAS_NetPredictionS=[]
      PSNR=[]
      SSIM=[]
      i=1 #Testing image number 1
      for CoverImage in mini_Batch_TestX:
            #Use Generator to generate stego images
            write_pgm((CoverImage*127.5+127.5),os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_CoverImages/{i}.pgm'))
            ProbabilityMap= generator(CoverImage,training=False)
            ModificationMap=TernaryEmbeddingSimulator(ProbabilityMap)
            StegoImage = tf.math.add(CoverImage, ModificationMap)
            write_pgm((StegoImage*127.5+127.5) , os.path.join(MasterPath,f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_StegoImages/{i}.pgm'))
            # # Compute security (for Stegnography) and Discrimability(for GAN) (Adverserial Accuracy)
            # Test ability to fool SRM+EC
            SRMPredictionC.append(SRM_EC(os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_CoverImages/{i}.pgm')))
            SRMPredictionS.append(SRM_EC(os.path.join(MasterPath, f'Evaluation\Quantitative\epoch {epoch + 1}\epoch_{epoch + 1}_StegoImages/{i}.pgm')))
            #Test ability to fool GBRAS_Net
            GBRAS_NetPredictionC.append(GBRAS_Net(CoverImage*127.5+127.5))
            GBRAS_NetPredictionS.append(GBRAS_Net(StegoImage*127.5+127.5))
            # Compute impercebtability (For stegnoography), and Overfitting, Perceptual Judgment,Sensitivity to Distortion,and Comp_& Sample effeciency  (For GAN) (Image Quality Measure)
            #Compute SSIM and PSNR
            SSIM.append(skimage.metrics.structural_similarity(tf.reshape(CoverImage,(256,256)).numpy(), tf.reshape(StegoImage,(256,256)).numpy()))
            PSNR.append(skimage.metrics.peak_signal_noise_ratio(tf.reshape(CoverImage,(256,256)).numpy(),tf.reshape(StegoImage,(256,256)).numpy()))
            i=i+1

      if (epoch==1):
          write_exel_file()
      # Compute avarage Detection error for SRM_EC steganalyzer
      SRMfalse_alarm = SRMPredictionC.count(+1.0) # number of time cover detected as stego
      SRMmissed_detection = SRMPredictionS.count(-1.0) # number of time stego detected as cover
      SRMDetection_Error = (SRMfalse_alarm + SRMmissed_detection) / ( len(SRMPredictionC) + len(SRMPredictionS))
      #Compute avarage Detection error for GBRAS_Net
      GBRAS_Netfalse_alarm = GBRAS_NetPredictionC.count(+1.0)  # number of time cover detected as stego
      GBRAS_Netmissed_detection = GBRAS_NetPredictionS.count(-1.0)  # number of time stego detected as cover
      GBRAS_NetDetection_Error = (GBRAS_Netfalse_alarm + GBRAS_Netmissed_detection) / (len(GBRAS_NetPredictionC) + len(GBRAS_NetPredictionC))
      #Computer avarage SSIM, and PSNR
      avarage_SSIM = sum(SSIM) / len(SSIM)
      avarage_PSNR = sum(PSNR) / len(PSNR)
      print('------------------------------------------------epoch',epoch+1,'-----------------------------------------------')
      print('SRM False alarm',SRMfalse_alarm)
      print('SRM missed_detection',SRMmissed_detection)
      print('SRMDetection_Error', SRMDetection_Error)
      print('GBRAS_Net False alarm',GBRAS_Netfalse_alarm)
      print('GBRAS_missed detection', GBRAS_Netmissed_detection)
      print('GBRAS_Net Detection_Error', GBRAS_NetDetection_Error )
      print('avarage_PSNR',avarage_PSNR)
      print('avarage_SSIM',avarage_SSIM)
      #---------------------------------------------------------------------------------------------------------
      # Append epoch results at the end of an Excel sheet
      #Create Dataframe with same column as excel sheet
      df= pd.DataFrame([[epoch,SRMfalse_alarm,SRMmissed_detection,SRMDetection_Error,GBRAS_Netfalse_alarm,GBRAS_Netmissed_detection,GBRAS_NetDetection_Error,avarage_PSNR,avarage_SSIM,epochTrainingTime]])
                       #columns=['Epoch#','GBRAS_Net False Alarm','GBRAS_Net Missed Detection','GBRAS_Net Detection error','Avarge PSNR','Avarage SSIM'])
      #append_df_to_excel(r'Evaluation/Quantitative/Results.xlsx', df, index=False)
      writer= pd.ExcelWriter('Evaluation/Quantitative/Results.xlsx',engine='openpyxl',mode='a')
      #try to open an existing workbook
      writer.book=openpyxl .load_workbook(r'Evaluation/Quantitative/Results.xlsx')
      # copy existing sheets
      writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
      # read existing file
      reader = pd.read_excel(r'Evaluation/Quantitative/Results.xlsx')
      # write out the new sheet
      df.to_excel(writer, index=False, header=False, startrow=len(reader)+1)
      writer.save()

      # Compute Overfitting
      # Compute Training Time
      # compute Capacity(for Stegnography)
      # avarage_SRMDetectionError,avarage_SSIM,avarage_PSNR

      return

def write_exel_file():
    from openpyxl.utils import get_column_letter
    #DataFrame
    # epoch#,SRM False Alarm, SRM Missed Detection, Avarage Detection error
    #GBRAS_Net False Alarm, GBRAS_Net missed Detection, GBRAS_Net Detection error,
    #avarge PSNR, SSIM
    df = pd.DataFrame({'Epoch#':[],
                       'SRM False Alarm': [],
                       'SRM Missed Detection': [],
                       'SRM Avarage Detection error': [],
                       'GBRAS_Net False Alarm':[],
                       'GBRAS_Net Missed Detection':[],
                       'GBRAS_Net Avarage Detection Error':[],
                       'Avarage PSNR':[],
                       'Avarage SSIM':[],
                       'Epoch Training Time':[]})

    #Create a pandas Ecel writer using XLsxWriter as engine
    writer= pd.ExcelWriter('Evaluation/Quantitative/Results.xlsx',engine='xlsxwriter',mode='w')

    #Convert the dataframe to an XlsxWriter Excel object
    df.to_excel(writer,sheet_name='Sheet1',index=False)
    #Resize column width
    writer.sheets['Sheet1'].set_column(1, 2, 20)
    writer.sheets['Sheet1'].set_column(3, 5,25)
    writer.sheets['Sheet1'].set_column(6, 6, 30)
    writer.sheets['Sheet1'].set_column(7, 9, 15)
    #add the boarder
    row_idx, col_idx = df.shape
    for r in range(row_idx):
        for c in range(col_idx):
            if c == 6:
                writer.write(r + 3, c, df.values[r, c],
                                writer.add_format({'border': 5, 'num_format': '0.00%'}))
            else:
                writer.write(r + 3, c, df.values[r, c],
                                writer.add_format({'border': 5, 'num_format': '0.00'}))

    #Save Workbook
    writer.save()



    '''
    global worksheet,row,workbook
    #"Evaluation/Quantitative/
    workbook = xlsxwriter.Workbook("Evaluation/Quantitative/results.xlsx")
    worksheet = workbook.add_worksheet()
    cell_format = workbook.add_format({'bold': True, 'font_color': 'red'})
    worksheet.write(0, 0, "epoch", cell_format)
    worksheet.write(0, 1, "SRM False Alarm", cell_format)
    worksheet.write(0, 2, "SRM Missed Detection", cell_format)
    worksheet.write(0, 3, "SRM Avarage Detection Error", cell_format)
    worksheet.write(0, 4, "GBRAS_Net False Alarm", cell_format)
    worksheet.write(0, 5, "GBRAS_Net Missed Detection", cell_format)
    worksheet.write(0, 6, "GBRAS_Net Avarage Detection Error", cell_format)
    worksheet.write(0, 7, "Avarage PSNR", cell_format)
    worksheet.write(0, 8, "Avarge SSIM", cell_format)
    row = 1'''
    return

def QualitativeEvaluation (epoch,generator,TestX):
      # Save the images generated during each epoch.
      os.makedirs(os.path.join(MasterPath, f'Evaluation\Qualitative\epoch {epoch+1}'), exist_ok=True)
      np.random.shuffle(TestX)
      mini_Batch_TestX = np.array_split(TestX, TestX.shape[0])
      TestImageNumber=1

      for image in mini_Batch_TestX:
            #Generating stego images
            CoverImage=image
            ProbabilityMap= generator(CoverImage,training=False)
            ModificationMap=TernaryEmbeddingSimulator(ProbabilityMap)
            StegoImage = tf.math.add(CoverImage, ModificationMap)
            #Plotting in Graph
            plt.subplot(2, 2,1,title='CoverImage')
            plt.imshow(tf.reshape(CoverImage, (256, 256))* 127.5 + 127.5,cmap='gray', vmin=0, vmax=255)
            plt.axis('off')
            plt.subplot(2, 2, 2,title='ProbabilityMap')
            plt.imshow(tf.reshape(ProbabilityMap, (256, 256))* 127.5 + 127.5,cmap='gray', vmin=0, vmax=255)
            plt.axis('off')
            plt.subplot(2, 2, 3,title='ModificationMap')
            plt.imshow(tf.reshape(ModificationMap, (256, 256))* 127.5 + 127.5,cmap='gray', vmin=0, vmax=255)
            plt.axis('off')
            plt.subplot(2, 2, 4,title='StegoImage')
            plt.imshow(tf.reshape(StegoImage, (256, 256))* 127.5 + 127.5, cmap='gray', vmin=0, vmax=255)
            plt.axis('off')
            plt.savefig(os.path.join(MasterPath, f'Evaluation\Qualitative\epoch {epoch+1}\ image {TestImageNumber}.png'))
            plt.close()
            TestImageNumber=TestImageNumber+1
      return



#Steganalayzer_Training_Testing_Data()
#SRM_EC_setup()


